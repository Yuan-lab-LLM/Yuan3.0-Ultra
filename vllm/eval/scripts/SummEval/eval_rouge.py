import json
from rouge_score import rouge_scorer
from tqdm import tqdm
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import argparse

def write_number_to_excel(
    file_path: str,
    value: float or int,
    cell: str = None,
    row: int = None,
    column: int = None,
    sheet_name: str = "Sheet1"
) -> None:
    if not isinstance(value, (int, float)):
        raise TypeError("▒~F~Y▒~E▒▒~Z~D▒~@▒▒~E须▒~X▒▒~U▒▒~U▒▒~H~V浮▒~B▒▒~U▒")

    if cell is None:
        if row is None or column is None:
            raise ValueError("请▒~L~G▒~Zcell▒~O~B▒~U▒▒~L▒~H~V▒~P~L▒~W▒▒~L~G▒~Zrow▒~R~Lcolumn▒~O~B▒~U▒")
    else:
        if row is not None or column is not None:
            raise ValueError("cell▒~O~B▒~U▒▒~Nrow/column▒~O~B▒~U▒▒~M▒~O▒▒~P~L▒~W▒使▒~T▒")

    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    # value= value*100
    if cell is not None:
        ws[cell] = value
    else:
        ws.cell(row=row, column=column, value=value)

    wb.save(file_path)
    print(f"已▒~F▒~U▒▒~W {value} ▒~F~Y▒~E▒ {file_path} ▒~Z~D {sheet_name} 工▒~\表▒~Z"
          f"{cell if cell else f'第{row}▒~L第{column}▒~H~W▒~H{get_column_letter(column)}{row}▒~I'}")

parser = argparse.ArgumentParser(description="QwQ-32B")
parser.add_argument("--path", type=str, default="../results/251104/Summary")
parser.add_argument("--excel-file", type=str, default="../results/251104/Summary/test.xlsx:summ1106")
parser.add_argument("--write-row", type=int, default=1)
args = parser.parse_args()
path = args.path
excel_file = args.excel_file.split(":")[0]
table_name = args.excel_file.split(":")[1]
write_row = args.write_row

data_names = ["cnn_dailymail", "xsum", "big_patent_10_percent", "billsum", "pubmed-summarization", "samsum", "wikihow"]
summary_keys = ["highlights", "summary", "abstract", "summary", "abstract", "summary", "summary"]
for data_name, summ_key in zip(data_names, summary_keys):
    '''
    data = []
    for file in os.listdir(path):
        file_path = os.path.join(path, file)
        with open(file_path, "r", encoding="utf-8") as f:
            examples = [json.loads(line) for line in f]
        data.extend(examples)
    '''
    with open(f"{path}/{data_name}.jsonl", "r", encoding="utf-8") as f:
        data = [json.loads(line) for line in f]
    scorer = rouge_scorer.RougeScorer(['rouge1', 'rouge2', 'rougeL'], use_stemmer=True)
    rouge_1, rouge_2, rouge_l = [], [], []
    for item in tqdm(data):
        answer = item['solution'].replace("<|end_of_sentence|>", "").replace("<eod>", "").replace("</think>", "")
        score = scorer.score(item[summ_key], answer)
        rouge_1.append(score['rouge1'].fmeasure)
        rouge_2.append(score['rouge2'].fmeasure)
        rouge_l.append(score['rougeL'].fmeasure)
    print(f"path: {path}")
    print(f"data: {data_name}")
    print(f"ROUGE-1: {sum(rouge_1)/len(rouge_1):.4f}")
    print(f"ROUGE-2: {sum(rouge_2)/len(rouge_2):.4f}")
    print(f"ROUGE-L: {sum(rouge_l)/len(rouge_l):.4f}")

    write_number_to_excel(file_path=excel_file, value=sum(rouge_1)/len(rouge_1), row=write_row, column=2, sheet_name=table_name)
    write_number_to_excel(file_path=excel_file, value=sum(rouge_2)/len(rouge_2), row=write_row, column=3, sheet_name=table_name)
    write_row += 1
